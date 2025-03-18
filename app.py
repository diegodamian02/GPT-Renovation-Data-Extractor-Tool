import argparse
import openpyxl
import docx
import PyPDF2
import json
import google.generativeai as genai
import re
from fuzzywuzzy import fuzz


# ✅ Set API Key Correctly
API_KEY = "USE GEMINI API"  # Replace with your actual API key
genai.configure(api_key=API_KEY)

# ✅ Extract Text from DOCX
def extract_text_from_doc(file):
    doc = docx.Document(file)
    return "\n".join([para.text for para in doc.paragraphs])

# ✅ Extract Text from PDF
def extract_text_from_pdf(file):
    text = ""
    with open(file, "rb") as pdf:
        reader = PyPDF2.PdfReader(pdf)
        for page in reader.pages:
            text += page.extract_text() + "\n"
    return text

# ✅ Extract Renovation Details (Better Budget Extraction & Fuzzy Matching)
def extract_renovation_details(text):
    model = "gemini-2.0-flash"
    prompt = f"""
You are a renovation expert helping extract structured data for a spreadsheet. 
Analyze the following conversation and extract only task-related details, including budgets.

Please return the output in **valid JSON format**, following this structure:
{{
    "task description 1": {{
        "CATEGORY": "GC / PLUMBING / ELECTRICAL / MILLWORK / SCOPE OF WORK", 
        "BUDGET": "XXXX", "PROPOSED": "", "COMMENT": "", "DWG REF": "", "LEAD": "Name (Speaker)"
    }},
    "task description 2": {{
        "CATEGORY": "...",
        "BUDGET": "...",
        "PROPOSED": "...",
        "COMMENT": "...",
        "DWG REF": "...",
        "LEAD": "..."
    }}
}}

Ensure:
- Extract **all valid budget values** (even if not preceded by "$", please analyze deeply numerical values to add into the spreadsheet).
- Convert **text-based budgets** into numerical values.
- Assign the **most relevant budget to the best matching description.**
- Default to `"BUDGET": "0"` if no budget is found or assign closest numerical value according to logic of the conversation.
- Extract valid external references (permit numbers, blueprint IDs) into **DWG REF**.
- Extract the **person in charge (LEAD)** from context. It cound be a speaker (e.g speaker 1 or speaker 2).
- Ensure the response is **valid JSON** (no markdown formatting).
- If it says it may cost a few hundred dollars convert it to any budget range 200-400
- 

Text to analyze:
---
{text}
"""

    try:
        model_instance = genai.GenerativeModel(model)
        response = model_instance.generate_content(prompt)

        # ✅ Debugging: Print Raw Response
        print("🔍 Gemini Raw Response:")
        print(response.text)

        if not response.text.strip():
            print("❌ Error: Empty response from Gemini AI.")
            return {}

        clean_response = response.text.strip().replace("```json", "").replace("```", "")
        extracted_data = json.loads(clean_response)

        # ✅ Improved Budget Matching (Better Context Assignment)
        budget_context_patterns = [
            r'(\d{3,})(?:\s+covers|, covers| all of that| includes| is about| is around)',
            r'costs?\s*(\d{3,})',
            r'around\s+(\d{3,})\s*-\s*(\d{3,})',  # Handles ranges
            r'(\d{3,})\s+is the estimate',
            r'budget\s+is\s+(\d{3,})',
            r'(\d{3,})\s+for the work'
        ]

        for task, details in extracted_data.items():
            task_text = f"{details.get('COMMENT', '')} {details.get('PROPOSED', '')}"

            budget_numbers = []

            # ✅ Extract budgets using regex patterns
            for pattern in budget_context_patterns:
                matches = re.findall(pattern, task_text)
                if matches:
                    for match in matches:
                        if isinstance(match, tuple):  # Handle budget ranges
                            start, end = map(int, match)
                            budget_numbers.append((start + end) / 2)
                        else:
                            budget_numbers.append(int(match))

            # ✅ Assign the best-matched budget
            details["BUDGET"] = str(int(sum(budget_numbers))) if budget_numbers else "0"

            # ✅ Extract DWG REF & LEAD
            ref_matches = re.findall(r'\b\d{4,}\b', task_text)
            if ref_matches:
                details["DWG REF"] = ", ".join(ref_matches)

            # ✅ Extract LEAD (Person in Charge)
            lead_match = re.search(r"(managed by|handled by|led by) (\w+ \w+)", task_text, re.IGNORECASE)
            details["LEAD"] = lead_match.group(2) if lead_match else "Unknown"

            # ✅ Fix PLUMBING category issue
            if details.get("CATEGORY") == "PLUMBING":
                details["PROPOSAL"] = details.pop("PROPOSED", False)

        return extracted_data

    except json.JSONDecodeError as json_err:
        print(f"❌ JSON Parsing Error: {json_err}")
        return {}

    except Exception as e:
        print(f"❌ Google Gemini API Error: {e}")
        return {}

def update_spreadsheet(extracted_data, spreadsheet):
    workbook = openpyxl.load_workbook(spreadsheet)

    sheet_columns = {
        "GC": ["BUDGET", "PROPOSED", "LEAD"],
        "PLUMBING": ["PROPOSAL"],
        "ELECTRICAL": ["BUDGET", "PROPOSED", "COMMENT", "DWG REF", "LEAD"],
        "MILLWORK": ["BUDGET", "PROPOSED", "COMMENT", "DWG REF", "LEAD"],
        "SCOPE OF WORK": ["BUDGET", "PROPOSED", "COMMENT", "DWG REF", "LEAD"],
        "PERMITS": ["BUDGET", "PROPOSED", "COMMENT", "DWG REF", "LEAD"]
    }

    for sheet in workbook.sheetnames:
        sheet_name = sheet.upper().strip()
        if sheet_name not in sheet_columns:
            print(f"⚠️ Sheet '{sheet_name}' not recognized, skipping...")
            continue

        print(f"📄 Processing Sheet: {sheet_name}")
        sheet_obj = workbook[sheet]
        relevant_tasks = {
            task.lower(): data for task, data in extracted_data.items()
            if data.get("CATEGORY", "").upper() == sheet_name
        }

        if not relevant_tasks:
            print(f"⚠️ No relevant data for {sheet_name}, skipping...")
            continue

        required_columns = sheet_columns[sheet_name]
        header_row_idx = None
        column_indices = {}

        # ✅ Locate header row with valid column names
        for row_idx, row in enumerate(sheet_obj.iter_rows(values_only=True), start=1):
            if row and any(col in row for col in required_columns):
                header_row_idx = row_idx
                column_indices = {col: row.index(col) for col in required_columns if col in row}
                break

        if not header_row_idx:
            print(f"⚠️ No header row found in {sheet_name}, skipping...")
            continue

        updated_tasks = set()

        # ✅ Update existing rows if a match is found
        for row in sheet_obj.iter_rows(min_row=header_row_idx + 1, max_row=sheet_obj.max_row):
            task_cell = row[0]
            if task_cell and task_cell.value:
                task_key = task_cell.value.strip().lower()

                # ✅ Find best match using fuzzy matching
                best_match, best_score = None, 0
                for existing_task in relevant_tasks.keys():
                    score = fuzz.partial_ratio(task_key, existing_task)
                    if score > best_score:
                        best_match, best_score = existing_task, score

                # ✅ If a strong match is found, update the existing row
                if best_match and best_score >= 40:
                    for col_name in required_columns:
                        if col_name in column_indices:
                            col_idx = column_indices[col_name]
                            cell = row[col_idx]
                            new_value = relevant_tasks[best_match].get(col_name, "")

                            # ✅ Convert budget ranges to averages
                            if col_name == "BUDGET" and "-" in new_value:
                                start, end = map(int, new_value.split("-"))
                                new_value = str((start + end) // 2)

                            # ✅ Force updating BUDGET even if 0
                            if col_name == "BUDGET" and (cell.value == "0" or not cell.value):
                                cell.value = new_value
                                print(f"💰 Updated: {best_match} -> {col_name}: {cell.value}")

                            # ✅ Force updating LEAD even if "Unknown"
                            elif col_name == "LEAD" and new_value not in ["Unknown", ""]:
                                cell.value = new_value
                                print(f"👤 Updated: {best_match} -> {col_name}: {cell.value}")

                            # ✅ Avoid unnecessary overwrites for other fields
                            elif cell.value not in [None, "", "0"] and cell.value == new_value:
                                continue

                            cell.value = new_value
                            updated_tasks.add(best_match)
                            print(f"✅ Updated: {best_match} -> {col_name}: {cell.value}")

        # ✅ Add missing tasks at the end of the table if no match was found
        for task_name, details in relevant_tasks.items():
            if task_name not in updated_tasks:
                new_row = [task_name] + [details.get(col, "") for col in required_columns]
                sheet_obj.append(new_row)
                print(f"➕ Added new row: {task_name} to {sheet_name}")

    workbook.save(spreadsheet)
    print(f"✅ Spreadsheet '{spreadsheet}' updated successfully!")


# ✅ Main Function
def main():
    parser = argparse.ArgumentParser(description="GPT Renovation Data Extractor Tool")
    parser.add_argument("file", help="Path to the transcript file (DOCX or PDF)")
    parser.add_argument("spreadsheet", help="Path to the Excel file to update")

    args = parser.parse_args()
    text = extract_text_from_doc(args.file) if args.file.endswith(".docx") else extract_text_from_pdf(args.file)

    print("📄 Extracting renovation details...")
    extracted_data = extract_renovation_details(text)

    print("📊 Updating spreadsheet...")
    update_spreadsheet(extracted_data, args.spreadsheet)
    print("🎉 Done!")

if __name__ == "__main__":
    main()
